"""
Combine per-respondent JSON files into a single CSV and Excel workbook.
"""

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR, JSON_DIR
from .schema import get_all_columns, MULTI_SELECT_EXPANSIONS


def _flatten_row(data: dict, source_file: str) -> dict:
    """Turn a raw JSON extraction into a flat dict ready for CSV export."""
    row: dict = {"source_file": source_file}

    for col in get_all_columns():
        if col == "source_file":
            continue
        val = data.get(col)
        if isinstance(val, list):
            row[col] = "; ".join(val) if val else ""
        elif val is None:
            row[col] = ""
        else:
            row[col] = val

    for array_col, options in MULTI_SELECT_EXPANSIONS.items():
        selected = set(data.get(array_col, []))
        for opt in options:
            safe = opt.lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace(",", "")
            flag_col = f"{array_col}__{safe}"
            row[flag_col] = 1 if opt in selected else 0

    return row


def _get_expanded_columns() -> list[str]:
    """Return ordered column names including binary-expanded multi-select flags."""
    cols = get_all_columns()
    expanded: list[str] = []
    for col in cols:
        expanded.append(col)
        if col in MULTI_SELECT_EXPANSIONS:
            for opt in MULTI_SELECT_EXPANSIONS[col]:
                safe = opt.lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace(",", "")
                expanded.append(f"{col}__{safe}")
    return expanded


def aggregate(json_dir: Path | None = None, output_dir: Path | None = None):
    """
    Read all JSON files in *json_dir* and write:
      - survey_results.csv
      - survey_results.xlsx
    into *output_dir*.
    """
    if json_dir is None:
        json_dir = JSON_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    json_files = sorted(json_dir.glob("*.json"))
    if not json_files:
        print("No JSON files found — nothing to aggregate.")
        return

    columns = _get_expanded_columns()
    rows: list[dict] = []

    for jf in json_files:
        data = json.loads(jf.read_text(encoding="utf-8"))
        source = jf.stem
        rows.append(_flatten_row(data, source))

    # --- CSV ---
    csv_path = output_dir / "survey_results.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    print(f"Wrote {csv_path}  ({len(rows)} rows)")

    # --- Excel ---
    xlsx_path = output_dir / "survey_results.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey Results"

    for ci, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = cell.font.copy(bold=True)

    for ri, row_data in enumerate(rows, start=2):
        for ci, col_name in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=row_data.get(col_name, ""))

    ws.freeze_panes = "B2"
    for ci in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    wb.save(str(xlsx_path))
    print(f"Wrote {xlsx_path}  ({len(rows)} rows)")
